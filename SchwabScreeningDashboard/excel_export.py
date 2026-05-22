"""
excel_export.py — Appends each momentum scan as a new dated sheet
to momentum_history.xlsx in the project folder.

Sheet layout per scan:
  Sheet 1 (or existing) "Index"  — run log: date, tickers scanned, top 5
  Sheet per scan        "YYYY-MM-DD HH:MM" — full ranked results + summary

Designed for backtesting: each scan is preserved forever, never overwritten.
"""

from __future__ import annotations

import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle

EXCEL_FILE = "momentum_history.xlsx"

# ── Colour palette ──────────────────────────────────────────────────────────
C_HEADER_BG  = "00274D"   # Schwab navy
C_HEADER_FG  = "FFFFFF"
C_SUBHDR_BG  = "E8EDF3"
C_SUBHDR_FG  = "1A2330"
C_STRONG_BG  = "E6F4EA"   # light green
C_MODERATE_BG= "E8F0FE"   # light blue
C_WARN_BG    = "FFF8E1"   # light amber
C_POSITIVE   = "1A6B35"
C_NEGATIVE   = "C0392B"
C_BORDER     = "DDE3EA"

# ── Column definitions ──────────────────────────────────────────────────────
COLUMNS = [
    # (header, df_field, width, number_format, description)
    ("Rank",               "rank",              6,  "0",          "Composite rank (1=best)"),
    ("Ticker",             "ticker",            9,  "@",          "Stock ticker symbol"),
    ("Price",              "current_price",    10,  '$#,##0.00',  "Last close price"),
    ("21d Return",         "ret_21",           11,  '0.0%',       "21-trading-day total return"),
    ("42d Return",         "ret_42",           11,  '0.0%',       "42-trading-day total return"),
    ("63d Return",         "ret_63",           11,  '0.0%',       "63-trading-day total return"),
    ("vs SPY 21d",         "vs_spy_21",        11,  '0.0%',       "Return minus SPY return (21d)"),
    ("vs SPY 63d",         "vs_spy_63",        11,  '0.0%',       "Return minus SPY return (63d)"),
    ("vs Sector 63d",      "vs_sector_63",     13,  '0.0%',       "Return minus sector ETF (63d)"),
    ("RS Slope",           "rs_slope_63",      10,  '0.0000',     "Slope of stock/SPY ratio (63d)"),
    ("Reg Slope (ann)",    "reg_slope",        13,  '0.00',       "Annualised linear regression slope of log price"),
    ("R-Squared",          "reg_r2",           10,  '0.00',       "R² of regression — trend consistency"),
    ("t-Statistic",        "reg_tstat",        10,  '0.0',        "t-stat of slope — statistical significance"),
    ("Volatility",         "vol_63",           10,  '0.0%',       "Annualised realised volatility (63d)"),
    ("Sharpe",             "sharpe_63",        10,  '0.00',       "Return / volatility (Sharpe-style)"),
    ("Sortino",            "sortino_63",       10,  '0.00',       "Return / downside volatility"),
    ("Max Drawdown",       "max_dd_63",        12,  '0.0%',       "Max peak-to-trough decline (63d)"),
    ("Calmar",             "calmar_63",        10,  '0.00',       "63d return / abs(max drawdown)"),
    ("RSI (14)",           "rsi_14",           10,  '0.0',        "14-day Relative Strength Index"),
    ("MACD",               "macd",             10,  "@",          "MACD signal: bullish / bearish"),
    ("Above MA20",         "above_ma20",        9,  "@",          "Price above 20-day moving average?"),
    ("Above MA50",         "above_ma50",        9,  "@",          "Price above 50-day moving average?"),
    ("MA20 > MA50",        "ma20_gt_ma50",      9,  "@",          "20-day MA above 50-day MA?"),
    ("% Up Days 63d",      "pct_up_63",        12,  '0.0%',       "Fraction of up-close days over 63d"),
    ("Vol / 20d Avg",      "vol_vs_20",        12,  '0.00',       "Today volume vs 20-day average"),
    ("Gap-Driven %",       "single_day_pct",   12,  '0.0%',       "Fraction of 63d log-return from single day"),
    ("Overextended",       "overextended",     12,  "@",          "Overextension flag (RSI/ATR/MA50)"),
    ("Flags",              "flags",            20,  "@",          "Specific overextension notes"),
    ("Composite Score",    "composite_score",  14,  '0.0',        "0-100 weighted momentum score"),
    ("Classification",     "classification",   18,  "@",          "Strong / Moderate / Weak / No Clear"),
]


def _thin_border():
    s = Side(border_style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _style_header(cell, bg=C_HEADER_BG, fg=C_HEADER_FG, size=10):
    cell.font      = Font(bold=True, color=fg, size=size, name="Arial")
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _thin_border()


def _style_data(cell, fmt="General", bold=False):
    cell.font      = Font(size=9, name="Arial", bold=bold)
    cell.alignment = Alignment(horizontal="right" if fmt not in ("@", "General") else "left",
                               vertical="center")
    cell.number_format = fmt
    cell.border    = _thin_border()


def _bool_display(val) -> str:
    if val is True or val == "True":  return "✓"
    if val is False or val == "False": return "✗"
    return str(val) if val is not None else ""


def _pct_or_nan(val) -> float | str:
    try:
        return float(val)
    except (TypeError, ValueError):
        return ""


# ── Index sheet ─────────────────────────────────────────────────────────────

def _ensure_index_sheet(wb: Workbook):
    if "Index" not in wb.sheetnames:
        ws = wb.create_sheet("Index", 0)
        ws.sheet_properties.tabColor = "00274D"
        headers = ["Scan Date", "Tickers Scanned", "Strong", "Moderate",
                   "Top 1", "Top 2", "Top 3", "Top 4", "Top 5", "Sheet Name"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(1, col, h)
            _style_header(c)
            ws.column_dimensions[get_column_letter(col)].width = 16
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 28


def _append_index_row(wb: Workbook, scan_dt: str, df: pd.DataFrame, sheet_name: str):
    ws = wb["Index"]
    strong   = int((df["classification"] == "Strong").sum())
    moderate = int((df["classification"] == "Moderate").sum())
    top5     = df.head(5)["ticker"].tolist()
    while len(top5) < 5:
        top5.append("")
    next_row = ws.max_row + 1
    row_data = [scan_dt, len(df), strong, moderate] + top5 + [sheet_name]
    for col, val in enumerate(row_data, 1):
        c = ws.cell(next_row, col, val)
        c.font      = Font(size=9, name="Arial")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _thin_border()
        if col == len(row_data):   # sheet name = hyperlink-style
            c.font = Font(size=9, name="Arial", color="0057A8", underline="single")


# ── Data sheet ───────────────────────────────────────────────────────────────

def _write_data_sheet(wb: Workbook, df: pd.DataFrame, sheet_name: str, scan_dt: str, oe_count: int = 0):
    ws = wb.create_sheet(sheet_name)

    # ── Title row ──
    ws.merge_cells("A1:AD1")
    title = ws["A1"]
    title.value     = f"Momentum Scan — {scan_dt}   |   {len(df)} stocks analysed   |   S&P 500 + Nasdaq 100"
    title.font      = Font(bold=True, size=12, name="Arial", color=C_HEADER_FG)
    title.fill      = PatternFill("solid", start_color=C_HEADER_BG)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Summary row ──
    strong   = int((df["classification"] == "Strong").sum())
    moderate = int((df["classification"] == "Moderate").sum())
    weak     = int((df["classification"] == "Weak/Unconfirmed").sum())
    ws.merge_cells("A2:AD2")
    summ = ws["A2"]
    summ.value = (f"Strong: {strong}   Moderate: {moderate}   Weak/Unconfirmed: {weak}   "
                  f"Overextended flags: {oe_count}   "
                  f"Gap-driven: {int((df['single_day_pct'] > 0.5).sum())}")
    summ.font      = Font(size=9, name="Arial", italic=True, color=C_SUBHDR_FG)
    summ.fill      = PatternFill("solid", start_color=C_SUBHDR_BG)
    summ.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    # ── Description row (tooltip-style) ──
    for col_idx, (hdr, field, width, fmt, desc) in enumerate(COLUMNS, 1):
        c = ws.cell(3, col_idx, desc)
        c.font      = Font(size=7, name="Arial", italic=True, color="6B7A8D")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.fill      = PatternFill("solid", start_color="F4F6F9")
        ws.row_dimensions[3].height = 28

    # ── Column headers ──
    for col_idx, (hdr, field, width, fmt, desc) in enumerate(COLUMNS, 1):
        c = ws.cell(4, col_idx, hdr)
        _style_header(c)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[4].height = 32
    ws.freeze_panes = "C5"    # freeze rank + ticker columns

    # ── Data rows ──
    DATA_START = 5
    for row_idx, (_, row) in enumerate(df.iterrows(), DATA_START):
        cls   = str(row.get("classification", ""))
        score = float(row.get("composite_score", 0) or 0)

        # Row background by classification
        if cls == "Strong":
            row_bg = C_STRONG_BG
        elif cls == "Moderate":
            row_bg = C_MODERATE_BG
        elif score >= 50:
            row_bg = C_WARN_BG
        else:
            row_bg = "FFFFFF"

        for col_idx, (hdr, field, width, fmt, desc) in enumerate(COLUMNS, 1):
            raw = row.get(field, "")

            # Type coercions
            if fmt == "@":
                if isinstance(raw, bool):
                    val = _bool_display(raw)
                else:
                    val = str(raw) if raw is not None and str(raw) != "nan" else ""
            elif "%" in fmt:
                val = _pct_or_nan(raw)
            else:
                try:
                    val = float(raw) if raw is not None and str(raw) not in ("nan", "") else ""
                except (TypeError, ValueError):
                    val = str(raw) if raw is not None else ""

            c = ws.cell(row_idx, col_idx, val)
            c.number_format = fmt
            c.font          = Font(size=9, name="Arial")
            c.fill          = PatternFill("solid", start_color=row_bg)
            c.border        = _thin_border()
            c.alignment     = Alignment(
                horizontal="right" if fmt not in ("@", "General") else "left",
                vertical="center"
            )

            # Inline colour for return columns
            if field in ("ret_21", "ret_42", "ret_63", "vs_spy_21", "vs_spy_63",
                         "vs_sector_63", "rs_slope_63", "reg_slope"):
                try:
                    num = float(raw)
                    c.font = Font(size=9, name="Arial",
                                  color=C_POSITIVE if num > 0 else (C_NEGATIVE if num < 0 else "000000"))
                except (TypeError, ValueError):
                    pass

            # Ticker bold
            if field == "ticker":
                c.font = Font(size=9, name="Arial", bold=True, color=C_HEADER_BG)

            # Score gradient colour (manually via font colour)
            if field == "composite_score":
                try:
                    s = float(raw)
                    if s >= 80:   clr = C_POSITIVE
                    elif s >= 65: clr = "0057A8"
                    elif s >= 50: clr = "E07B00"
                    else:         clr = C_NEGATIVE
                    c.font = Font(size=9, name="Arial", bold=True, color=clr)
                except (TypeError, ValueError):
                    pass

        ws.row_dimensions[row_idx].height = 15

    # ── Autofilter on header row ──
    last_col = get_column_letter(len(COLUMNS))
    last_row = DATA_START + len(df) - 1
    ws.auto_filter.ref = f"A4:{last_col}4"

    return ws


# ── Summary sheet ────────────────────────────────────────────────────────────

def _write_summary_sheet(wb: Workbook, df: pd.DataFrame, scan_dt: str, sheet_name: str = ""):
    base  = f"Summ {scan_dt[:10]}"
    sname = base if base not in wb.sheetnames else f"Summ {sheet_name}"
    ws = wb.create_sheet(sname)
    ws.sheet_properties.tabColor = "00A651"

    def section(row, title):
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row, 1, title)
        c.font  = Font(bold=True, size=10, name="Arial", color=C_HEADER_FG)
        c.fill  = PatternFill("solid", start_color=C_HEADER_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        return row + 1

    def mini_table(ws, start_row, sub_df, extra_cols=None):
        cols = ["rank", "ticker", "composite_score", "classification",
                "ret_63", "vs_spy_63"]
        if extra_cols:
            cols += extra_cols
        headers = ["Rank", "Ticker", "Score", "Class", "63d Ret", "vs SPY"] + \
                  ([c.replace("_", " ").title() for c in (extra_cols or [])])
        for ci, h in enumerate(headers, 1):
            c = ws.cell(start_row, ci, h)
            _style_header(c, bg="1A3A5C", size=9)
            ws.column_dimensions[get_column_letter(ci)].width = 14
        r = start_row + 1
        for _, row in sub_df.iterrows():
            for ci, col in enumerate(cols, 1):
                val = row.get(col, "")
                if isinstance(val, float):
                    val = round(val, 4)
                c = ws.cell(r, ci, val)
                c.font      = Font(size=9, name="Arial")
                c.alignment = Alignment(horizontal="center")
                c.border    = _thin_border()
            r += 1
        return r + 1

    r = 1
    ws.merge_cells("A1:F1")
    c = ws.cell(1, 1, f"MOMENTUM SCAN SUMMARY — {scan_dt}")
    c.font = Font(bold=True, size=13, name="Arial", color=C_HEADER_FG)
    c.fill = PatternFill("solid", start_color=C_HEADER_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26
    r = 2

    r = section(r, "🏆 Top 10 by Composite Score")
    r = mini_table(ws, r, df.head(10))

    r = section(r, "🎯 Top 5 Risk-Adjusted (Sharpe)")
    r = mini_table(ws, r, df.sort_values("sharpe_63", ascending=False).head(5), ["sharpe_63", "sortino_63"])

    r = section(r, "💪 Top 5 vs SPY (63d)")
    r = mini_table(ws, r, df.sort_values("vs_spy_63", ascending=False).head(5), ["vs_spy_63", "vs_sector_63"])

    r = section(r, "📈 Top 5 Trend Persistence (R²)")
    r = mini_table(ws, r, df[df["reg_slope"] > 0].sort_values("reg_r2", ascending=False).head(5), ["reg_r2", "reg_tstat"])

    oe = df[df["overextended"].isin([True, "✓"])]
    r = section(r, f"⚠ Overextended ({len(oe)} flagged)")
    if not oe.empty:
        r = mini_table(ws, r, oe.head(10), ["rsi_14", "flags"])

    gd = df[pd.to_numeric(df["single_day_pct"], errors="coerce") > 0.5]
    r = section(r, f"🚫 Gap-Driven / Rejected ({len(gd)})")
    if not gd.empty:
        r = mini_table(ws, r, gd.head(10), ["single_day_pct"])

    return ws


# ── Public API ───────────────────────────────────────────────────────────────

def append_scan(df: pd.DataFrame, filepath: str = EXCEL_FILE) -> str:
    """
    Append a momentum scan to the persistent Excel workbook.

    Parameters:
        df:       DataFrame from momentum.run_screen()
        filepath: Path to the Excel file (created if it doesn't exist)

    Returns:
        Absolute path of the saved file.
    """
    if df is None or df.empty:
        raise ValueError("No scan results to export")

    now        = datetime.now()
    scan_dt    = now.strftime("%Y-%m-%d %H:%M")
    sheet_name = now.strftime("%Y-%m-%d %H%M")   # no colon — safe for Excel

    df = df.copy()
    # Count stats BEFORE converting booleans to symbols
    _oe_count = int(df["overextended"].apply(lambda x: x is True or x == True or str(x) == "True").sum()) if "overextended" in df.columns else 0
    # Clean boolean columns for Excel display
    for col in ("above_ma20", "above_ma50", "ma20_gt_ma50", "overextended"):
        if col in df.columns:
            df[col] = df[col].apply(_bool_display)

    # Load or create workbook
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
    else:
        wb = Workbook()
        # Remove default blank sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    _ensure_index_sheet(wb)
    _write_data_sheet(wb, df, sheet_name, scan_dt, oe_count=_oe_count)
    _write_summary_sheet(wb, df, scan_dt, sheet_name)
    _append_index_row(wb, scan_dt, df, sheet_name)

    # Keep Index sheet first
    if "Index" in wb.sheetnames:
        wb.move_sheet("Index", offset=-len(wb.sheetnames) + 1)

    wb.save(filepath)
    return os.path.abspath(filepath)
